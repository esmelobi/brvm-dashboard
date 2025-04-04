import os
import io
import re
import fitz
import pandas as pd
from datetime import datetime
from collections import defaultdict
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# 🧡 Titres à surveiller en priorité
FAVORIS = ["ORANGE COTE D'IVOIRE (ORAC)", "SAPH CI (SAPH)", "SONATEL SN (SNTS)"]
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
FOLDER_ID = '1w2W-SI19l3qgpJKOEGCIiKS2fOIwx3OY'  # <-- à personnaliser
BULLETIN_DIR = 'bulletins'
DATA_FILE = 'data/portefeuille.csv'

# 🔐 Authentification
def authenticate_drive():
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    else:
        flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
        creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('drive', 'v3', credentials=creds)

# 📥 Téléchargement
def download_bulletins():
    os.makedirs(BULLETIN_DIR, exist_ok=True)
    service = authenticate_drive()
    query = f"'{FOLDER_ID}' in parents and mimeType='application/pdf'"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    for file in results.get('files', []):
        name = file['name']
        path = os.path.join(BULLETIN_DIR, name)
        if not os.path.exists(path):
            request = service.files().get_media(fileId=file['id'])
            with open(path, 'wb') as f:
                downloader = MediaIoBaseDownload(f, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
            print(f"⬇️ Téléchargé : {name}")

# 📄 Extraction des hausses/baisses
def extract_top_movers_from_pdf(path, date_str):
    with fitz.open(path) as doc:
        text = "\n".join(page.get_text() for page in doc)

    data = []
    i = 0
    lines = text.splitlines()
    section = None
    while i < len(lines):
        line = lines[i].strip()
        if "PLUS FORTES HAUSSES" in line:
            section = "hausse"
            i += 4
        elif "PLUS FORTES BAISSES" in line:
            section = "baisse"
            i += 4
        elif section and line:
            try:
                titre = lines[i].strip()
                # Si le titre ressemble à un montant, on saute
                if re.match(r'^\d[\d\s]*$', titre) or "%" in titre or len(titre) < 4:
                    i += 1
                    continue
                cours = int(lines[i+1].replace(" ", "").replace("FCFA", ""))
                var_jour = float(lines[i+2].replace("%", "").replace(",", "."))
                var_annuelle = float(lines[i+3].replace("%", "").replace(",", "."))
                data.append({
                    'date': date_str,
                    'titre': titre,
                    'cours': cours,
                    'variation_jour': var_jour,
                    'variation_annuelle': var_annuelle,
                    'type': section
                })
                i += 4
            except:
                i += 1
        else:
            i += 1
    return data

# 📊 Génération du tableau de bord
def update_portfolio():
    os.makedirs('data', exist_ok=True)
    all_data = []
    for file in sorted(os.listdir(BULLETIN_DIR)):
        if file.endswith(".pdf"):
            date_match = re.search(r'\d{4}-\d{2}-\d{2}', file)
            if not date_match:
                continue
            date_str = date_match.group(0)
            path = os.path.join(BULLETIN_DIR, file)
            all_data.extend(extract_top_movers_from_pdf(path, date_str))

    df = pd.DataFrame(all_data)
    df.to_csv(DATA_FILE, index=False)

    stats = defaultdict(lambda: {'hausses': 0, 'baisses': 0, 'total_var': 0.0, 'last_var': 0.0, 'last_date': ''})
    for _, row in df.iterrows():
        t = row['titre']
        if row['type'] == 'hausse':
            stats[t]['hausses'] += 1
        else:
            stats[t]['baisses'] += 1
        stats[t]['total_var'] += row['variation_jour']
        if row['date'] > stats[t]['last_date']:
            stats[t]['last_var'] = row['variation_jour']
            stats[t]['last_date'] = row['date']

    portfolio = []
    for titre, st in stats.items():
        if st['hausses'] >= 3 and st['total_var'] > 5:
            reco = '🟢 Achat'
        elif st['baisses'] >= 3 and st['total_var'] < -5:
            reco = '🔴 Vente'
        else:
            reco = '🟡 Observer'
        portfolio.append({
            'Titre': titre,
            'Jours en Hausse': st['hausses'],
            'Jours en Baisse': st['baisses'],
            'Variation Totale (%)': round(st['total_var'], 2),
            'Dernière Variation (%)': round(st['last_var'], 2),
            'Recommandation': reco
        })

    df_final = pd.DataFrame(portfolio)
    export_excel(df_final)
    df_final.to_csv('data/recommandations.csv', index=False)
    df_final = df_final.sort_values(by='Variation Totale (%)', ascending=False)
    print("\n📋 RECOMMANDATIONS DU JOUR\n")
    print(df_final.to_string(index=False))
    export_excel(df_final)

def export_excel(df_final):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()

    # Onglet principal : Recommandations
    ws1 = wb.active
    ws1.title = "Recommandations"
    fills = {
        '🟢 Achat': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        '🔴 Vente': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        '🟡 Observer': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    }

    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
            elif c_idx == 6:
                fill = fills.get(value, None)
                if fill:
                    cell.fill = fill

    for col in ws1.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max_length + 2

    # Onglet Favoris
    ws2 = wb.create_sheet("Favoris")
    df_fav = df_final[df_final['Titre'].isin(FAVORIS)]
    for r_idx, row in enumerate(dataframe_to_rows(df_fav, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)

    for col in ws2.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max_length + 2

    # Graphique : Jours en Hausse vs Baisse
    chart = BarChart()
    chart.title = "Favoris – Hausses vs Baisses"
    chart.y_axis.title = "Nombre de jours"
    chart.x_axis.title = "Titre"
    data_ref = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=ws2.max_row)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, f"H2")

    # Enregistrer
    wb.save("data/recommandations.xlsx")
    print("📈 Excel enrichi généré : data/recommandations.xlsx")


# 🚀 Exécution
if __name__ == "__main__":
    download_bulletins()
    update_portfolio()
